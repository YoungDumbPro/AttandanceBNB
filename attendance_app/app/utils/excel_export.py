"""Excel report generation using openpyxl.

Generates .xlsx reports for employee attendance data with
formatted columns and summary statistics.
"""
from io import BytesIO
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.timezone_helpers import utc_to_local


def generate_excel_report(employee, records, emp_tz):
    """Generate an Excel attendance report for an employee.
    
    Args:
        employee: Employee model instance.
        records: List of Attendance records.
        emp_tz: ZoneInfo timezone for the employee.
    
    Returns:
        BytesIO object containing the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance - {employee.username}"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(size=11, color="666666")
    summary_font = Font(bold=True, size=11)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title Section
    ws['A1'] = f"Attendance Report - {employee.username}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Timezone: {employee.timezone}"
    ws['A2'].font = subtitle_font
    ws.merge_cells('A2:D2')
    
    ws['A3'] = f"Country: {employee.country} | Role: {employee.role}"
    ws['A3'].font = subtitle_font
    ws.merge_cells('A3:D3')

    # Header Row (Row 5)
    headers = ['Date', 'Check In', 'Check Out', 'Hours Worked']
    header_row = 5
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data Rows
    total_hours = 0.0
    days_worked = set()
    current_row = header_row + 1
    
    for record in records:
        # Convert to local timezone for display
        local_checkin = utc_to_local(record.check_in_utc, employee.timezone)
        local_checkout = utc_to_local(record.check_out_utc, employee.timezone)
        
        # Date
        date_cell = ws.cell(row=current_row, column=1, value=local_checkin.strftime('%Y-%m-%d'))
        date_cell.border = thin_border
        
        # Check In
        checkin_cell = ws.cell(row=current_row, column=2, value=local_checkin.strftime('%I:%M %p'))
        checkin_cell.border = thin_border
        checkin_cell.alignment = Alignment(horizontal='center')
        
        # Check Out
        if record.check_out_utc:
            checkout_str = local_checkout.strftime('%I:%M %p')
        else:
            checkout_str = 'Still Active'
        checkout_cell = ws.cell(row=current_row, column=3, value=checkout_str)
        checkout_cell.border = thin_border
        checkout_cell.alignment = Alignment(horizontal='center')
        
        # Hours Worked
        hours = record.worked_hours
        hours_cell = ws.cell(row=current_row, column=4, value=round(hours, 2))
        hours_cell.border = thin_border
        hours_cell.alignment = Alignment(horizontal='center')
        hours_cell.number_format = '0.00'
        
        # Accumulate stats
        total_hours += hours
        days_worked.add(local_checkin.date())
        
        current_row += 1

    # Summary Section
    summary_row = current_row + 2
    num_days = len(days_worked)
    avg_hours = round(total_hours / num_days, 2) if num_days > 0 else 0
    
    ws.cell(row=summary_row, column=1, value='Summary').font = summary_font
    
    ws.cell(row=summary_row + 1, column=1, value='Total Hours:')
    ws.cell(row=summary_row + 1, column=2, value=round(total_hours, 2))
    ws.cell(row=summary_row + 1, column=2).number_format = '0.00'
    
    ws.cell(row=summary_row + 2, column=1, value='Days Worked:')
    ws.cell(row=summary_row + 2, column=2, value=num_days)
    
    ws.cell(row=summary_row + 3, column=1, value='Average Hours/Day:')
    ws.cell(row=summary_row + 3, column=2, value=avg_hours)
    ws.cell(row=summary_row + 3, column=2).number_format = '0.00'

    # Set column widths
    column_widths = [15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
